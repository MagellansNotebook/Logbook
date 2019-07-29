from collections import Counter
from config_file import database
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from peewee import *
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
from tkinter.ttk import Progressbar
from tkinter import ttk
import datetime
import time
import random
import tkinter as tk
import threading
import os
################################################################
#  							       #
#         Main tk App class to display the window box          #
#                                                              #
################################################################
class App(tk.Tk):
	def __init__(self, *args, **kwargs):
		"""Main Window Display"""
		tk.Tk.__init__(self, *args, **kwargs)
		self.window = tk.Frame(self)
		self.window.pack(expand=True, fill='both', side='top')
		#<---------------------------------------------------------------------------| Menu Functions
		menu = tk.Menu(self.window)
		# Menu
		file = Menu(menu)
		edit = Menu(menu)
		view = Menu(menu)
		# File Menu
		menu.add_cascade(label="File", menu=file)
		file.add_command(label="Information", command=lambda: self.showFrame(InfoPage))
		file.add_command(label="Scan", command=lambda: self.showFrame(ScanPage))
		file.add_command(label="Track", command=lambda: self.showFrame(TrackPage))
		file.add_command(label="Upload", command=lambda: UploadFunc("Message", "Loading File..."))
		file.add_command(label="Exit", command=self.window.quit)
		# Edit Menu
		menu.add_cascade(label="Edit", menu=edit)
		edit.add_command(label="Add", command=lambda: self.showFrame(AddPage))
		edit.add_command(label="Edit", command=lambda: self.showFrame(EditPage))
		edit.add_command(label="Delete", command=lambda: self.showFrame(DeletePage))
		# View Menu
		menu.add_cascade(label="View", menu=view)
		view.add_command(label="Search Item", command=lambda: self.showFrame(SearchPage))
		tk.Tk.config(self, menu=menu)
		#<---------------------------------------------------------------------------| Blank Page
		self.blank = {}
		blank = BlankPage(self.window, self)
		self.blank[BlankPage] = blank
		blank.grid(row=0, column=0, sticky='nw')
		#<---------------------------------------------------------------------------| Window Frames Classes
		self.frame = {}
		for win in (InfoPage, AddPage, DeletePage, EditPage, ScanPage, SearchPage, TrackPage):
			frame = win(self.window, self)
			self.frame[win] = frame
			frame.grid(row=0, column=0, sticky='nw')
		self.showFrame(InfoPage)
	def showBlank(self, BlankPage):
		"""Uses a Blank Page as a background"""
		blank = self.blank[BlankPage]
		blank.tkraise()
	def showFrame(self, win):
		"""Moves a Window frame on top of group pile"""
		self.showBlank(BlankPage)
		frame = self.frame[win]
		frame.tkraise()
################################################################
#  							       #
#                     Displays Treeview                        #
#                                                              #
################################################################
class DisplayTreeView(ttk.Treeview):
	def __init__(self, labelFrame, page):
		"""Display's Treeview"""
		ttk.Treeview.__init__(self, labelFrame)
		#<---------------------------------------------------------------------------| TreeView column name
		columnName = ('Location', 'Type', 'NSN', 'Model', 'Description', 
			'Serial Number', 'Asset Number', 'SCA Number', 'Tracking Type', 'Quantity', 'Date Last Seen') 
		#<---------------------------------------------------------------------------| TreeView height
		if page == "Delete Equipment" or page == "Edit Equipment" or page == "Add Equipment":
			self.tree = ttk.Treeview(labelFrame, columns=columnName, height=20)
		elif page == "Search Equipment":
			self.tree = ttk.Treeview(labelFrame, columns=columnName, height=25)
		elif page == "Track Equipment":
			self.tree = ttk.Treeview(labelFrame, columns=columnName, height=15)
		elif page == "TrackRecord Equipment":
			self.tree = ttk.Treeview(labelFrame, columns=columnName, height=10)
		#<---------------------------------------------------------------------------| TreeView heading name
		if page == "Delete Equipment" or page == "Edit Equipment" or page == "Add Equipment" or page == "Track Equipment":
			self.tree.heading('#0', text="UiD")
			self.tree.heading('#1', text="Location", anchor='w')
			self.tree.heading('#2', text="Type", anchor='w')
			self.tree.heading('#3', text="NSN", anchor='w')
			self.tree.heading('#4', text="Model", anchor='w')
			self.tree.heading('#5', text="Description", anchor='w')
			self.tree.heading('#6', text="Serial Number", anchor='w')
			self.tree.heading('#7', text="Asset Number", anchor='w')
			self.tree.heading('#8', text="SCA Number", anchor='w')
			self.tree.heading('#9', text="Tracking Type", anchor='w')
			self.tree.heading('#10', text="Quantity", anchor='n')
			# Treeview column size
			self.tree.column('#0', stretch=tk.NO, minwidth=40, width=40)
			self.tree.column('#1', stretch=tk.NO, minwidth=220, width=220)
			self.tree.column('#2', stretch=tk.NO, minwidth=110, width=110)
			self.tree.column('#3', stretch=tk.NO, minwidth=110, width=110)
			self.tree.column('#4', stretch=tk.NO, minwidth=102, width=102)
			self.tree.column('#5', stretch=tk.NO, minwidth=306, width=306)
			self.tree.column('#6', stretch=tk.NO, minwidth=130, width=130)
			self.tree.column('#7', stretch=tk.NO, minwidth=110, width=110)
			self.tree.column('#8', stretch=tk.NO, minwidth=110, width=110)
			self.tree.column('#9', stretch=tk.NO, minwidth=110, width=110)
			self.tree.column('#10', stretch=tk.NO, minwidth=75, width=75, anchor='n')
			# Treeview grid position
			self.tree.grid(row=0, columnspan=4, sticky='nsew')
		elif page == "Search Equipment" or page == "TrackRecord Equipment":
			self.tree.heading('#0', text="UiD", command=lambda: self.treeSortUniqueID(self.tree, "#0", False))
			self.tree.heading('#1', text="Location", command=lambda: self.treeSortCol(self.tree, "Location", False), anchor='w')
			self.tree.heading('#2', text="Type", command=lambda: self.treeSortCol(self.tree, "Type", False), anchor='w')
			self.tree.heading('#3', text="NSN", command=lambda: self.treeSortCol(self.tree, "NSN", False), anchor='w')
			self.tree.heading('#4', text="Model", command=lambda: self.treeSortCol(self.tree, "Model", False), anchor='w')
			self.tree.heading('#5', text="Description", command=lambda: self.treeSortCol(self.tree, "Description", False), anchor='w')
			self.tree.heading('#6', text="Serial Number", command=lambda: self.treeSortCol(self.tree, "Serial Number", False), anchor='w')
			self.tree.heading('#7', text="Asset Number", command=lambda: self.treeSortCol(self.tree, "Asset Number", False), anchor='w')
			self.tree.heading('#8', text="SCA Number", command=lambda: self.treeSortCol(self.tree, "SCA Number", False), anchor='w')
			self.tree.heading('#9', text="Tracking Type", command=lambda: self.treeSortCol(self.tree, "Tracking Type", False), anchor='w')
			self.tree.heading('#10', text="Quantity", command=lambda: self.treeSortCol(self.tree, "Quantity", False), anchor='n')
			self.tree.heading('#11', text="Date Last Seen", command=lambda: self.treeSortCol(self.tree, "Date Last Seen", False), anchor='w')
			# Treeview column size
			self.tree.column('#0', stretch=tk.NO, minwidth=50, width=50)
			self.tree.column('#1', stretch=tk.NO, minwidth=202, width=202)
			self.tree.column('#2', stretch=tk.NO, minwidth=90, width=90)
			self.tree.column('#3', stretch=tk.NO, minwidth=88, width=88)
			self.tree.column('#4', stretch=tk.NO, minwidth=88, width=88)
			self.tree.column('#5', stretch=tk.NO, minwidth=255, width=255)
			self.tree.column('#6', stretch=tk.NO, minwidth=180, width=180)
			self.tree.column('#7', stretch=tk.NO, minwidth=105, width=105)
			self.tree.column('#8', stretch=tk.NO, minwidth=100, width=100)
			self.tree.column('#9', stretch=tk.NO, minwidth=100, width=100)
			self.tree.column('#10', stretch=tk.NO, minwidth=72, width=72, anchor='n')
			self.tree.column('#11', stretch=tk.NO, minwidth=104, width=104)
			# Treeview grid position
			self.tree.grid(row=0, columnspan=4, sticky='nsew')
		#<---------------------------------------------------------------------------| Treeview Scroll Bar
		if page == "Delete Equipment" or page == "Edit Equipment" or page == "Add Equipment":
			yScrollbar = ttk.Scrollbar(labelFrame, orient='vertical')
			yScrollbar.place(x=1423, y=1, height=425)
			self.tree.configure(yscrollcommand=yScrollbar.set)
			yScrollbar.configure(command=self.tree.yview)
		elif page == "Track Equipment":
			yScrollbar = ttk.Scrollbar(labelFrame, orient='vertical')
			yScrollbar.place(x=1423, y=1, height=325)
			self.tree.configure(yscrollcommand=yScrollbar.set)
			yScrollbar.configure(command=self.tree.yview)
		elif page == "Search Equipment":
			yScrollbar = ttk.Scrollbar(labelFrame, orient='vertical')
			yScrollbar.place(x=1420, y=1, height=525)
			self.tree.configure(yscrollcommand=yScrollbar.set)
			yScrollbar.configure(command=self.tree.yview)
	#<---------------------------------------------------------------------------| Treeview Functions
	def clearTreeviewButton(self):
		"""Removes display in Treeview"""
		for entry in self.tree.get_children():
			self.tree.delete(entry)
	def displayTree(self,query):
		"""Displays entry in the Treeview"""
		for entry in query:
			timestamp = entry.scaEList.timestampMaster.strftime('%d-%b-%y')
			self.tree.insert('','end',text=str(entry.id),value=(
				entry.scaEList.itemLoc, 
				entry.scaEList.itemTyp,
				entry.scaEList.itemNSN, 
				entry.scaEList.itemMod, 
				entry.scaEList.itemDes, 
				entry.scaEList.itemSer, 
				entry.scaEList.itemAss, 
				entry.scaNum, 
				entry.scaEList.itemTrk,
				entry.scaEList.itemQty, 
				timestamp,))
	def treeSortUniqueID(self, tv, col, reverse):
		"""Sorts UniqueID Ascending/Descending"""
		l = [(tv.item(k)['text'], k) for k in tv.get_children()]
		l.sort(key=lambda t: t[0], reverse=reverse)
		for index, (val, k) in enumerate(l):
			tv.move(k, '', index)
		tv.heading(col, command=lambda: self.treeSortUniqueID(tv, col, not reverse))
	def treeSortCol(self, tv, col, reverse):
		"""Sorts Values Ascending/Descending"""
		l = [(tv.set(k, col), k) for k in tv.get_children('')]
		l.sort(reverse=reverse)
		for index, (val, k) in enumerate(l):
			tv.move(k, '', index)
		tv.heading(col, command=lambda: self.treeSortCol(tv, col, not reverse))
	def treeView(self):
		"""Passess in Treeview"""
		return self.tree
################################################################
#  							       #
#                     Message Box Classes                      #
#                                                              #
################################################################
class MessageBox(tk.Toplevel):
	def __init__(self, title, message):
		"""Message Box for errors"""
		super().__init__()
		self.details_expanded = False
		self.title(title)
		self.geometry('400x100+{}+{}'.format(self.master.winfo_x(), self.master.winfo_y()))
		self.resizable(False, False)
		self.rowconfigure(0, weight=0)
		self.rowconfigure(1, weight=1)
		self.columnconfigure(0, weight=0)
		self.columnconfigure(1, weight=1)
		#<---------------------------------------------------------------------------| Message displayed on the MessageBox
		tk.Label(self, text=message).grid(row=0, column=0, columnspan=3, pady=(7,7), padx=(7,7), sticky='ew')
		#<---------------------------------------------------------------------------| Button
		tk.Button(self, text="Ok", command=self.destroy, width=20).grid(row=1, column=1,  padx=7, sticky='e')
class UploadFunc(tk.Toplevel):
	def __init__(self, title, message):
		"""Displays a message box with a progress bar when uploading a file into the database"""
		#<---------------------------------------------------------------------------| Opens File dialog box
		self.filename = filedialog.askopenfilename(title="Select File", filetypes=(("Excel Files", "*.xlsx"),("All Files","*.*")))
		# Creates a message box
		super().__init__()
		self.details_expanded = False
		self.title(title)
		self.geometry('400x100+{}+{}'.format(self.master.winfo_x(), self.master.winfo_y()))
		self.resizable(False, False)
		self.rowconfigure(0, weight=0)
		self.rowconfigure(1, weight=1)
		self.rowconfigure(2, weight=1)
		self.columnconfigure(0, weight=0)
		self.columnconfigure(1, weight=1)
		#<---------------------------------------------------------------------------| Label
		tk.Label(self, text=message).grid(row=0, column=0, columnspan=3, pady=10, padx=10, sticky='ew')
		#<---------------------------------------------------------------------------| Button
		self.button = Button(self, text="Ok", state='disabled', command=self.destroy, width=15)
		self.button.grid(row=2, column=1,  padx=15, sticky='ew')
		#<---------------------------------------------------------------------------| Sets the progress bar
		self.progress = ttk.Progressbar(self, orient='horizontal', length=100, mode='determinate')
		self.progressBar()
	def progressBar(self):
		"""Progress Bar when uploading an excel file"""
		def progress():
			self.progress.grid(row=1, column=1, padx=15, sticky='ew')
			self.progress.start()
			msg = self.upload()
			self.progress.stop()
			self.progress.grid_forget()
			self.button.config(state='normal')
			if not self.filename or msg == "Error":
			# Displays message if self.filename is blank close the message box
				self.destroy()
			else:
				tk.Label(self, text="Upload Complete").grid(row=0, column=0, columnspan=3, pady=10, padx=10, sticky='ew')
		threading.Thread(target=progress).start()
	def upload(self):
		"""Uploads Excel file to the database"""
		try:
			if self.filename:
				wb = load_workbook(self.filename)
				ws = wb["Sheet1"]
				# Place holders for list and counters
				progressCount = 0
				for row in ws.iter_rows(min_row=2, max_col=12, values_only=True):
					# Assings a variable 
					loc = row[0]
					typ = row[1]
					nsn = row[2]
					mod = row[3]
					des = row[4]
					ser = row[5]
					ass = row[6]
					sca = row[7]
					trk = row[8]
					qty = row[9]
					dte = row[10]
					# Generates Random number
					num = ""
					for i in range(0,9):
						num += str(random.randint(0,9))
					#<---------------------------------------------------------------------------| Assigns random number if serial/asset num is blank
					if not ser:
						ser = "<tmp-"+num+">"
					elif not ass:
						ass = "<tmp-"+num+">"
					# add value into the database
					errorMsg = database.fileUpload(loc, typ, nsn, mod, des, ser, ass, sca, trk, qty, dte)
					if not errorMsg:
						# Number of items to be added to the database 
						progressCount =+ progressCount + 1
						os.system('CLS')
						print(progressCount, " > Number of entries in the eTracker Database")
					else:
						messagebox.showerror("Error", "An error occured during the upload process.\nThe date column must be formatted to date in the Excel Spreadsheet")
						return errorMsg
		#<---------------------------------------------------------------------------| Error Messages			
		except KeyError:
			messagebox.showerror("Error", "An error occured during the upload process.\n Tab title is not set to 'Sheet1'")
			errorMsg = "Error"
			return errorMsg
		except:
			messagebox.showerror("Error", "An error occured during the upload process. Check data in the Excel Spreadsheet")
			errorMsg = "Error"
			return errorMsg
################################################################
#  							       #
#                    Page Class - Add Page                     #
#                                                              #
################################################################
class AddPage(tk.Frame):
	def __init__(self, parent, child):
		"""Add Page"""
		tk.Frame.__init__(self, parent)
		#<---------------------------------------------------------------------------| LabelFrame for Add Equipment
		labelFrameAdd = LabelFrame(self, text="Add Equipment", font=style)
		labelFrameAdd.pack(expand=1, fill='both')
		# Labels
		tk.Label(labelFrameAdd, text="Location").grid(row=2, column=0, sticky='w')
		tk.Label(labelFrameAdd, text="Type").grid(row=2, column=1, sticky='w')
		tk.Label(labelFrameAdd, text="NSN").grid(row=2, column=2, sticky='w')
		tk.Label(labelFrameAdd, text="Model").grid(row=2, column=3, sticky='w')
		tk.Label(labelFrameAdd, text="Description").grid(row=2, column=4, sticky='w')
		tk.Label(labelFrameAdd, text="Serial Number").grid(row=2, column=5, sticky='w')
		tk.Label(labelFrameAdd, text="Asset Number").grid(row=2, column=6, sticky='w')
		tk.Label(labelFrameAdd, text="SCA Number").grid(row=2, column=7, sticky='w')
		tk.Label(labelFrameAdd, text="Tracking Type").grid(row=2, column=8, sticky='w')
		tk.Label(labelFrameAdd, text="Quantity").grid(row=2, column=9, sticky='w')
		#<---------------------------------------------------------------------------| Dictionary
		# Dict for Entry
		self.columnA = {}
		self.columnB = {}
		self.columnC = {}
		self.columnD = {}
		self.columnE = {}
		self.columnF = {}
		self.columnG = {}
		self.columnH = {}
		self.columnI = {}
		self.columnJ = {}
		# Dict for String Variable
		self.cellLoc = {}
		self.cellType = {}
		self.cellNSN = {}
		self.cellModel = {}
		self.cellDes = {}
		self.cellSerial = {}
		self.cellAsset = {}
		self.cellSCA = {}
		self.cellTrk = {}
		self.cellQty = {}
		#<---------------------------------------------------------------------------| String Variable
		# Assigning String Variable to Text Variable
		for i in range(0,10):
			self.cellLoc[i] = tk.StringVar()
			self.cellType[i] = tk.StringVar()
			self.cellNSN[i] = tk.StringVar()
			self.cellModel[i] = tk.StringVar()
			self.cellDes[i] = tk.StringVar()
			self.cellSerial[i] = tk.StringVar()
			self.cellAsset[i] = tk.StringVar()
			self.cellSCA[i] = tk.StringVar()
			self.cellTrk[i] = tk.StringVar()
			self.cellQty[i] = tk.StringVar()
		#<---------------------------------------------------------------------------| Creating rows and columns
		for i in range(0,10):
			self.columnA[i] = Entry(labelFrameAdd, textvariable=self.cellLoc[i], width=35)
			self.columnA[i].grid(row=4 + i, column=0, sticky='w')
			self.columnB[i] = Entry(labelFrameAdd, textvariable=self.cellType[i], width=20)
			self.columnB[i].grid(row=4 + i, column=1, sticky='w')
			self.columnC[i] = Entry(labelFrameAdd, textvariable=self.cellNSN[i], width=15)
			self.columnC[i].grid(row=4 + i, column=2, sticky='w')
			self.columnD[i] = Entry(labelFrameAdd, textvariable=self.cellModel[i], width=22)
			self.columnD[i].grid(row=4 + i, column=3, sticky='w')
			self.columnE[i] = Entry(labelFrameAdd, textvariable=self.cellDes[i], width=60)
			self.columnE[i].grid(row=4 + i, column=4, sticky='w')
			self.columnF[i] = Entry(labelFrameAdd, textvariable=self.cellSerial[i], width=25)
			self.columnF[i].grid(row=4 + i, column=5, sticky='w')
			self.columnG[i] = Entry(labelFrameAdd, textvariable=self.cellAsset[i], width=15)
			self.columnG[i].grid(row=4 + i, column=6, sticky='w')
			self.columnH[i] = Entry(labelFrameAdd, textvariable=self.cellSCA[i], width=16)
			self.columnH[i].grid(row=4 + i, column=7, sticky='w')
			self.columnI[i] = Entry(labelFrameAdd, textvariable=self.cellTrk[i], width=15)
			self.columnI[i].grid(row=4 + i, column=8, sticky='w')
			self.columnJ[i] = Entry(labelFrameAdd, textvariable=self.cellQty[i], width=10)
			self.columnJ[i].grid(row=4 + i, column=9, sticky='w')
		#<---------------------------------------------------------------------------| LabelFrame for Treeview Add Equipment
		labelFrameTree = LabelFrame(self, text="Equipment List", font=style)
		labelFrameTree.pack(expand=1, fill='both')
		self.displayTree = DisplayTreeView(labelFrameTree, "Add Equipment")
		#<---------------------------------------------------------------------------| Passes in Treeview from the Class DisplayTreeView
		self.treeview = self.displayTree.treeView()
		self.treeview.bind('<Double-1>', self.deleteValue)
		#<---------------------------------------------------------------------------| LabelFrame for Action Buttons
		labelFrameAction = LabelFrame(self, text="Action", font=style)
		labelFrameAction.pack(expand=1, fill='both')
		#<---------------------------------------------------------------------------| Buttons
		tk.Button(labelFrameAction, text="Add", command=self.addButton, width=20).grid(row=1, column=0, pady=5, sticky='w')
		tk.Button(labelFrameAction, text="Upload", command=self.uploadButton, width=20).grid(row=1, column=1, pady=5, sticky='w')
		tk.Button(labelFrameAction, text="Clear Add List", command=self.clearAddListButton, width=20).grid(row=1, column=2, pady=5, sticky='w')
		tk.Button(labelFrameAction, text="Clear Equipment List", command=self.displayTree.clearTreeviewButton, width=20).grid(row=1, column=3, pady=5, sticky='w')
	def addButton(self):
		"""Add entry value to Treeview"""
		for i in range(0,9):
			#<---------------------------------------------------------------------------| Get String Variable
			loc = self.columnA[i].get()
			typ = self.columnB[i].get()
			nsn = self.columnC[i].get()
			mod = self.columnD[i].get()
			des = self.columnE[i].get()
			ser = self.columnF[i].get()
			ass = self.columnG[i].get()
			sca = self.columnH[i].get()
			trk = self.columnI[i].get()
			qty = self.columnJ[i].get()
			# Generate randow number for equipment without serial or asset number
			num = ""
			for i in range(0,9):
				num += str(random.randint(0,9))
			# Assigning random number for equipment with no serial or asset number
			if not ser:
				ser = "<tmp-"+num+">"
				if loc and typ and nsn and mod and des and sca and trk and qty:
					self.treeview.insert('','end', text=">", value=(loc, typ, nsn, mod, des, ser, ass, sca, trk, qty))
			elif not ass:
				ass = "<tmp-"+num+">"
				if loc and typ and nsn and mod and des and sca and trk and qty:
					self.treeview.insert('','end', text=">", value=(loc, typ, nsn, mod, des, ser, ass, sca, trk, qty))
			else:
				if loc and typ and nsn and mod and des and sca and trk and qty:
					self.treeview.insert('','end', text=">", value=(loc, typ, nsn, mod, des, ser, ass, sca, trk, qty))
		# Clears entry list
		self.clearAddListButton()
	def uploadButton(self):
		"""Add equipment to Database"""
		entryList = []
		for entry in self.treeview.get_children():
			for values in self.treeview.item(entry, 'values'):
				entryList.append(values)
		# List to check for duplicated values		
		dupList = []
		# Creates a maximum range from the treeview
		lengthOfList = len(entryList)//10
		# Checks for any Serial and Asset number duplicates
		try:
			for i in range(0,lengthOfList+1):
				ser = entryList[i*10+5]
				ass = entryList[i*10+6]
				dupList.append(ser)
				dupList.append(ass)
		except IndexError:
			dupVal = [k for k, v in Counter(dupList).items() if v > 1]
			if dupVal:
				# If a duplicate is found then a messagebox will appear
				msgBox = MessageBox("Duplicate Value", dupVal)
			else:
				try:
					date = datetime.datetime.now()
					for i in range(0,lengthOfList+1):
						loc = entryList[i*10]
						typ = entryList[i*10+1]
						nsn = entryList[i*10+2]
						mod = entryList[i*10+3]
						des = entryList[i*10+4]
						ser = entryList[i*10+5]
						ass = entryList[i*10+6]
						sca = entryList[i*10+7]
						trk = entryList[i*10+8]
						qty = entryList[i*10+9]
						# Passess in the variable to the Database
						database.addValue(loc, typ, nsn, mod, des, ser, ass, sca, trk, qty, date)
						# Removes entry in Treeview
						self.displayTree.clearTreeviewButton()
				except IndexError:
					pass
	def clearAddListButton(self):
		"""Clears entry values"""
		for i in range(0,9):
			self.columnA[i].delete(first=0, last=100)
			self.columnB[i].delete(first=0, last=100)
			self.columnC[i].delete(first=0, last=100)
			self.columnD[i].delete(first=0, last=100)
			self.columnE[i].delete(first=0, last=100)
			self.columnF[i].delete(first=0, last=100)
			self.columnG[i].delete(first=0, last=100)
			self.columnH[i].delete(first=0, last=100)
			self.columnI[i].delete(first=0, last=100)
			self.columnJ[i].delete(first=0, last=100)
	def deleteValue(self, event):
		"""delete a string value into an entry box. Executes on double click command"""
		try:
			# Highlights the selected row and deletes it
			currentItem = self.treeview.focus()
			self.treeview.delete(currentItem)
		except IndexError:
			# Ignores error
			pass
################################################################
#  							       #
#                 Page Class - Blank Page                      #
#                                                              #
################################################################
class BlankPage(tk.Frame):
	def __init__(self, parent, child):
		"""Blank Page"""
		tk.Frame.__init__(self, parent)
		labelPage = LabelFrame(self, width=1450, height=780)
		labelPage.pack(expand=0, fill='both')
################################################################
#  							       #
#                Page Class - Delete Page                      #
#                                                              #
################################################################
class DeletePage(tk.Frame):
	def __init__(self, parent, child):
		"""Delete entry from the Database"""
		tk.Frame.__init__(self, parent)
		#<---------------------------------------------------------------------------| LabelFrame for Delete Equipment
		labelFrameDel = LabelFrame(self, text="Delete Equipment", font=style)
		labelFrameDel.pack(expand=1, fill='both')
		#<---------------------------------------------------------------------------| String Value
		self.idValue = tk.StringVar()
		self.serValue = tk.StringVar()
		self.assValue = tk.StringVar()
		#<---------------------------------------------------------------------------| Labels and Entry Value
		# Unique ID
		tk.Label(labelFrameDel, text="Unique ID").grid(row=2, column=0, sticky='w')
		self.entryID = Entry(labelFrameDel, textvariable=self.idValue, width=20)
		self.entryID.bind('<Return>', self.addButton)
		self.entryID.grid(row=4, column=0, sticky='w')
		# Serial Number
		tk.Label(labelFrameDel, text="Serial Number").grid(row=2, column=1, padx=10, sticky='w')
		self.entrySer = Entry(labelFrameDel, textvariable=self.serValue, width=30)
		self.entrySer.bind('<Return>', self.addButton)
		self.entrySer.grid(row=4, column=1, padx=10, pady=5, sticky='w')
		# Asset Number
		tk.Label(labelFrameDel, text="Asset Number").grid(row=2, column=2, sticky='w')
		self.entryAss = Entry(labelFrameDel, textvariable=self.assValue, width=30)
		self.entryAss.bind('<Return>', self.addButton)
		self.entryAss.grid(row=4, column=2, sticky='w')
		#<---------------------------------------------------------------------------| Buttons
		tk.Button(labelFrameDel, text="Add", command=self.addButton, width=20).grid(row=4, column=3, padx=10, sticky='w')
		tk.Button(labelFrameDel, text="Clear Search Item", command=self.clearSearchButton, width=20).grid(row=4, column=4, sticky='w')
		#<---------------------------------------------------------------------------| LabelFrame for Treeview and Display Treeview
		labelFrameTree = LabelFrame(self, text="Equipment List", font=style)
		labelFrameTree.pack(expand=1, fill='both')
		self.displayTree = DisplayTreeView(labelFrameTree, "Delete Equipment")
		#<---------------------------------------------------------------------------| Passes in Treeview from the Class DisplayTreeView
		self.treeview = self.displayTree.treeView()
		#<---------------------------------------------------------------------------| LabelFrame for Action Buttons
		labelFrameAction = LabelFrame(self, text="Action", font=style)
		labelFrameAction.pack(expand=1, fill='both')
		#<---------------------------------------------------------------------------| Buttons
		tk.Button(labelFrameAction, text="Commit", command=self.deleteButton, width=20).grid(column=0, row=10, padx=10, pady=10, sticky='w')
		tk.Button(labelFrameAction, text="Cancel", command=self.displayTree.clearTreeviewButton, width=20).grid(column=1, row=10, padx=10, pady=10, sticky='w')
	def addButton(self, event=None):
		"""Add value into a treeview. Executes to enter command"""
		idValue = self.idValue.get()
		serValue = self.serValue.get()
		assValue = self.assValue.get()
		#<---------------------------------------------------------------------------| Displays values into Treeview
		if idValue:
			query = database.searchID(idValue)
		elif serValue:
			query = database.searchSer(serValue)
		elif assValue:
			query = database.searchAss(assValue)
		# Displays value in treeview and clears entry box
		self.displayTree.displayTree(query)	
		self.clearSearchButton()
	def deleteButton(self):
		"""Deletes entry fromt the database"""
		entryList = []
		for entry in self.treeview.get_children():
			for values in self.treeview.item(entry, 'values'):
				entryList.append(values)
		# Creates a maximum range from the treeview
		lengthOfList = len(entryList)//11
		try:
			for i in range(0,lengthOfList+1):
				ser = entryList[i*11+5]
				database.delete(ser)
				self.displayTree.clearTreeviewButton()
		except IndexError:
			pass
	def clearSearchButton(self):
		"""Clears value in entry"""
		listEntry = (self.entryID, self.entrySer, self.entryAss)
		for cells in listEntry:
			cells.delete(first=0, last=100)
################################################################
#  							       #
#                  Page Class - Edit Page                      #
#                                                              #
################################################################
class EditPage(tk.Frame):
	def __init__(self, parent, child):
		"""Edit entry from the database"""
		tk.Frame.__init__(self, parent)
		#<---------------------------------------------------------------------------| LabelFrame for Edit Equipment
		labelFrameEdit = LabelFrame(self, text="Edit Equipment", font=style)
		labelFrameEdit.pack(expand=1, fill='both')
		#<---------------------------------------------------------------------------| String Value
		self.idValue = tk.StringVar()
		self.serValue = tk.StringVar()
		self.assValue = tk.StringVar()
		#<---------------------------------------------------------------------------| Labels and Entry Value
		# Unique ID
		tk.Label(labelFrameEdit, text="Unique ID").grid(row=2, column=0, sticky='w')
		self.entryID = Entry(labelFrameEdit, textvariable=self.idValue, width=20)
		self.entryID.bind('<Return>', self.addButton)
		self.entryID.grid(row=4, column=0, sticky='w')
		# Serial Number
		tk.Label(labelFrameEdit, text="Serial Number").grid(row=2, column=1, padx=10, sticky='w')
		self.entrySer = Entry(labelFrameEdit, textvariable=self.serValue, width=30)
		self.entrySer.bind('<Return>', self.addButton)
		self.entrySer.grid(row=4, column=1, padx=10, pady=5, sticky='w')
		# Asset Number
		tk.Label(labelFrameEdit, text="Asset Number").grid(row=2, column=2, sticky='w')
		self.entryAss = Entry(labelFrameEdit, textvariable=self.assValue, width=30)
		self.entryAss.bind('<Return>', self.addButton)
		self.entryAss.grid(row=4, column=2, sticky='w')
		#<---------------------------------------------------------------------------| LabelFrame for Treeview and Display Treeview
		labelFrameTree = LabelFrame(self, text="Equipment List", font=style)
		labelFrameTree.pack(expand=1, fill='both')
		self.displayTree = DisplayTreeView(labelFrameTree, "Edit Equipment")
		#<---------------------------------------------------------------------------| Insert Value into an entry using double click function
		self.treeview = self.displayTree.treeView()
		self.treeview.bind('<Double-1>', self.insertValue)
		#<---------------------------------------------------------------------------| Button
		tk.Button(labelFrameEdit, text="Add", command=self.addButton, width=20).grid(row=4, column=3, padx=10, sticky='w')
		tk.Button(labelFrameEdit, text="Clear Search Item", command=self.clearSearchButton, width=20).grid(row=4, column=4, sticky='w')
		tk.Button(labelFrameEdit, text="Clear Edit List", command=self.displayTree.clearTreeviewButton, width=20).grid(row=4, column=5, padx=10, sticky='w')
		#<---------------------------------------------------------------------------| LabelFrame for Entry Box
		labelFrameEditBox = LabelFrame(self, text="Edit Item", font=style)
		labelFrameEditBox.pack(expand=1, fill='both')
		#<---------------------------------------------------------------------------| String Values
		self.valuesID = 0
		self.stringVal = {}
		for i in range(0,10):
			self.stringVal[i] = tk.StringVar()
		#<---------------------------------------------------------------------------| Label and Entry Box
		# Location
		tk.Label(labelFrameEditBox, text="Location").grid(row=2, column=0, sticky='w')
		self.entryLoc = Entry(labelFrameEditBox, textvariable=self.stringVal[0], width=30)
		self.entryLoc.grid(row=4, column=0, sticky='w')
		# Type
		tk.Label(labelFrameEditBox, text="Type").grid(row=2, column=1, sticky='w')
		self.entryTyp = Entry(labelFrameEditBox, textvariable=self.stringVal[1], width=18)
		self.entryTyp.grid(row=4, column=1, sticky='w')
		# NSN
		tk.Label(labelFrameEditBox, text="NSN").grid(row=2, column=2, sticky='w')
		self.entryNSN = Entry(labelFrameEditBox, textvariable=self.stringVal[2], width=20)
		self.entryNSN.grid(row=4, column=2, sticky='w')
		# Model
		tk.Label(labelFrameEditBox, text="Model").grid(row=2, column=3, sticky='w')
		self.entryMod = Entry(labelFrameEditBox, textvariable=self.stringVal[3], width=20)
		self.entryMod.grid(row=4, column=3, sticky='w')
		# Descriptor
		tk.Label(labelFrameEditBox, text="Description").grid(row=2, column=4, sticky='w')
		self.entryDes = Entry(labelFrameEditBox, textvariable=self.stringVal[4], width=45)
		self.entryDes.grid(row=4, column=4, sticky='w')
		# Serail Number
		tk.Label(labelFrameEditBox, text="Serial Number").grid(row=2, column=5, sticky='w')
		self.entrySer = Entry(labelFrameEditBox, textvariable=self.stringVal[5], width=30)
		self.entrySer.grid(row=4, column=5, sticky='w')
		# Asset Number
		tk.Label(labelFrameEditBox, text="Asset Number").grid(row=2, column=6, sticky='w')
		self.entryAss = Entry(labelFrameEditBox, textvariable=self.stringVal[6], width=20)
		self.entryAss.grid(row=4, column=6, sticky='w')
		# SCA Number
		tk.Label(labelFrameEditBox, text="SCA Number").grid(row=2, column=7, sticky='w')
		self.entrySCA = Entry(labelFrameEditBox, textvariable=self.stringVal[7], width=20)
		self.entrySCA.grid(row=4, column=7, sticky='w')
		# Tracking Type
		tk.Label(labelFrameEditBox, text="Tracking Type").grid(row=2, column=8, sticky='w')
		self.entryTrk = Entry(labelFrameEditBox, textvariable=self.stringVal[8], width=20)
		self.entryTrk.grid(row=4, column=8, sticky='w')
		# Quantity
		tk.Label(labelFrameEditBox, text="Quantity").grid(row=2, column=9, sticky='w')
		self.entryQty = Entry(labelFrameEditBox, textvariable=self.stringVal[9], width=10)
		self.entryQty.grid(row=4, column=9, sticky='w')
		#<---------------------------------------------------------------------------| LabelFrame for Button
		labelFrameAction = LabelFrame(self, text="Action", font=style)
		labelFrameAction.pack(expand=1, fill='both')
		#<---------------------------------------------------------------------------| Button
		tk.Button(labelFrameAction, text="Commit", command=self.commitButton, width=20).grid(row=10, column=0, padx=10, pady=10, sticky='w')
		tk.Button(labelFrameAction, text="Cancel", command=self.cancelButton, width=20).grid(row=10, column=1, padx=10, pady=10, sticky='w')
	def addButton(self, event=None):
		"""Add value into a treeview. Executes to enter command"""
		idValue = self.idValue.get()
		serValue = self.serValue.get()
		assValue = self.assValue.get()
		#<---------------------------------------------------------------------------| Displays values into Treeview
		if idValue:
			query = database.searchID(idValue)
		elif serValue:
			query = database.searchSer(serValue)
		elif assValue:
			query = database.searchAss(assValue)
		# Displays value in treeview and clears entry box
		self.displayTree.displayTree(query)		
		self.clearSearchButton()
	def cancelButton(self):
		"""Removes the values in the Entry Box"""
		self.entryLoc.delete(first=0, last=100)
		self.entryTyp.delete(first=0, last=100)
		self.entryNSN.delete(first=0, last=100)
		self.entryMod.delete(first=0, last=100)
		self.entryDes.delete(first=0, last=100)
		self.entrySer.delete(first=0, last=100)
		self.entryAss.delete(first=0, last=100)
		self.entrySCA.delete(first=0, last=100)
		self.entryTrk.delete(first=0, last=100)
		self.entryQty.delete(first=0, last=100)
	def clearSearchButton(self):
		"""Removes the values in the Entry Box"""
		listEntry = (self.entryID, self.entrySer, self.entryAss)
		for cells in listEntry:
			cells.delete(first=0, last=100)
	def commitButton(self):
		"""Create new entry in the Database"""
		#<---------------------------------------------------------------------------| Get String Variable
		ID = self.valuesID
		loc = self.stringVal[0].get()
		typ = self.stringVal[1].get()
		nsn = self.stringVal[2].get()
		mod = self.stringVal[3].get()
		des = self.stringVal[4].get()
		ser = self.stringVal[5].get()
		ass = self.stringVal[6].get()
		sca = self.stringVal[7].get()
		trk = self.stringVal[8].get()
		qty = self.stringVal[9].get()
		# Generate randow number for equipment without serial or asset number
		num = ""
		for i in range(0,9):
			num += str(random.randint(0,9))
		# Assigning random number for equipment with no serial or asset number
		if loc and typ and nsn and mod and des and ser and ass and sca and trk:
			database.update(ID, loc, typ, nsn, mod, des, ser, ass, sca, trk, qty, dte=None)
		elif not ser:
			ser = "<tmp-"+num+">"
			database.update(ID, loc, typ, nsn, mod, des, ser, ass, sca, trk, qty, dte=None)
		elif not ass:
			ass = "<tmp-"+num+">"
			database.update(ID, loc, typ, nsn, mod, des, ser, ass, sca, trk, qty, dte=None)
		# Clears entry list
		self.cancelButton()
	def insertValue(self, event):
		"""Inserts a string value into an entry box. Executes on double click command"""
		try:
			# Highlights the selected row
			currentItem = self.treeview.focus()
			contents = self.treeview.item(currentItem)
			values = contents['values']
			# Insert tree values into entry box
			self.entryLoc.insert(END, values[0])
			self.entryTyp.insert(END, values[1])
			self.entryNSN.insert(END, values[2])
			self.entryMod.insert(END, values[3])
			self.entryDes.insert(END, values[4])
			self.entrySer.insert(END, values[5])
			self.entryAss.insert(END, values[6])
			self.entrySCA.insert(END, values[7])
			self.entryTrk.insert(END, values[8])
			self.entryQty.insert(END, values[9])
			# Gets the ID number of the treeview content
			contentsID = self.treeview.item(currentItem)
			valuesID = contentsID['text']
			self.valuesID = valuesID
			# Removes the row from the treeview table
			delCurrentItem = self.treeview.focus()
			delContents = self.treeview.delete(delCurrentItem)
		except IndexError:
			pass
################################################################
#  						               #
#                  Page Class - Info Page                      #
#                                                              #
################################################################
class InfoPage(tk.Frame):
	def __init__(self, parent, child):
		"""Scan and creates a track record for an entry value"""
		tk.Frame.__init__(self, parent)
		#<---------------------------------------------------------------------------| LabelFrame for Information
		labelFrameDate = LabelFrame(self, text="Information", font=style)
		labelFrameDate.pack(expand=1, fill='both')
		# Label for the display date
		self.date = "Date: " + datetime.datetime.now().strftime('%d-%b-%y')
		tk.Label(labelFrameDate, text=self.date).grid(row=0, column=0, padx=10, sticky='nw')
		#<---------------------------------------------------------------------------| Button
		tk.Button(labelFrameDate, text="Refresh", command=self.action, width=20).grid(row=1, column=0, padx=10, pady=10, sticky='w')
		#<---------------------------------------------------------------------------| LabelFrame for Treeview
		labelFrame = LabelFrame(self, text="Statistics", font=style)
		labelFrame.pack(expand=1, fill='both')
		# Treeview 
		columnNameStatistics = ('#1', '#2', '#3', '#4', '#5', '#6', '#7') 
		self.tree = ttk.Treeview(labelFrame, columns=columnNameStatistics, height=12)
		self.tree.heading('#0', text="")
		self.tree.heading('#1', text="", anchor='w')
		self.tree.heading('#2', text="Total", anchor='n')
		self.tree.heading('#3', text="Serial Track", anchor='n')
		self.tree.heading('#4', text="Qty Track", anchor='n')
		self.tree.heading('#5', text="", anchor='n')
		self.tree.heading('#6', text="", anchor='n')
		self.tree.heading('#7', text="", anchor='n')
		# Treeview column size
		self.tree.column('#0', stretch=tk.NO, minwidth=0, width=0)
		self.tree.column('#1', stretch=tk.NO, minwidth=280, width=280)
		self.tree.column('#2', stretch=tk.NO, minwidth=100, width=100, anchor='n')
		self.tree.column('#3', stretch=tk.NO, minwidth=100, width=100, anchor='n')
		self.tree.column('#4', stretch=tk.NO, minwidth=100, width=100, anchor='n')
		self.tree.column('#5', stretch=tk.NO, minwidth=180, width=180, anchor='n')
		self.tree.column('#6', stretch=tk.NO, minwidth=180, width=180, anchor='n')
		self.tree.column('#7', stretch=tk.NO, minwidth=150, width=150, anchor='n')
		# Grid position
		self.tree.grid(row=0, column=0, columnspan=1, sticky='nsew')
		# Insert defualt value
		self.tree.insert('','end', text='', value=("Number of Equipment checked in the last 30 days", 0, 0, 0))
		self.tree.insert('','end', text='', value=("Number of Equipment unchecked in the last 30 days", 0, 0, 0))
		self.tree.insert('','end', text='', value=("Total Number of Equipment", 0, 0, 0))
		# Blank Space
		self.tree.insert('','end', text='', value=(""))
		# Column Headers
		self.tree.insert('','end', text='', value=("SCA Accounts", "Numbers", "Serial Track", "Qty Track", "Number of Equipment Check", 
			"Number of Equipment uncheck", "Current Date Check"))
		# Display SCA entries
		query = database.searchSCANum()
		for entry in query:
			self.tree.insert('','end', text='', value=(entry.scaNum, 0, 0, 0, 0, 0, 0))
		#<---------------------------------------------------------------------------| LabelFrame for Treeview
		labelFrameNSN = LabelFrame(self, text="NSN Information", font=style)
		labelFrameNSN.pack(expand=1, fill='both')
		#<---------------------------------------------------------------------------| Button
		tk.Button(labelFrameNSN, text="Display", command=self.showNSN, width=20).grid(row=0, column=0, padx=10, pady=10, sticky='w')
		#<---------------------------------------------------------------------------| Copy Option
		tk.Label(labelFrameNSN, text="Copy and Paste value into the Search Page - 'NSN'", font=style).grid(row=2, column=0, padx=10, pady=5, sticky='w')
		self.entryNSN = Entry(labelFrameNSN, textvariable="", width=55)
		self.entryNSN.grid(row=3, column=0, padx=10, pady=5, sticky='w')
		# Treeview
		columnNameNSN = ('#1', '#2', '#3') 
		self.treeNSN = ttk.Treeview(labelFrameNSN, columns=columnNameNSN, height=10)
		self.treeNSN.heading('#0', text="")
		self.treeNSN.heading('#1', text="Description", command=lambda: self.treeSortCol(self.treeNSN, "#1", False), anchor='w')
		self.treeNSN.heading('#2', text="Type", command=lambda: self.treeSortCol(self.treeNSN, "#2", False), anchor='w')
		self.treeNSN.heading('#3', text="NSN", command=lambda: self.treeSortCol(self.treeNSN, "#3", False), anchor='w')
		# Treeview column size
		self.treeNSN.column('#0', stretch=tk.NO, minwidth=0, width=0)
		self.treeNSN.column('#1', stretch=tk.NO, minwidth=310, width=310)
		self.treeNSN.column('#2', stretch=tk.NO, minwidth=120, width=120, anchor='w')
		self.treeNSN.column('#3', stretch=tk.NO, minwidth=130, width=130, anchor='w')
		# Bind command to the selected item
		self.treeNSN.bind('<Double-1>', self.selectItem)
		# Grid position
		self.treeNSN.grid(row=1, column=0, columnspan=1, sticky='nsew')
		# Scroll 
		yScrollbar = ttk.Scrollbar(labelFrameNSN, orient='vertical')
		yScrollbar.place(x=545, y=47, height=225)
		self.treeNSN.configure(yscrollcommand=yScrollbar.set)
		yScrollbar.configure(command=self.treeNSN.yview)
	def action(self):
		"""Inserts the current value into the treeview"""
		# Clear Values
		self.clear()
		# Current Counts
		self.tree.insert('','end', text='', value=("Number of Equipment checked in the last 30 days ", 
			database.countBetweenDates(), database.countBetSer(), database.countBetQty()))
		self.tree.insert('','end', text='', value=("Number of Equipment unchecked in the last 30 days", database.countTotalUncheck(), 
			database.countTotalSerUncheck(), database.countTotalAssUncheck()))
		self.tree.insert('','end', text='', value=("Total Number of Equipment", 
			database.count(), database.countSerialTrack(), database.countQtyTrack()))
		self.tree.insert('','end', text='', value=(""))
		self.tree.insert('','end', text='', value=("SCA Accounts", "Numbers", "Serial Track", "Qty Track", "Number of Equipment Check", 
			"Number of Equipment uncheck", "Current Date Check"))
		# Display sca entries
		query = database.searchSCANum()
		for entry in query:
			self.tree.insert('','end', text='', value=(entry.scaNum, database.countSCA(entry.scaNum), database.countBetSCASer(entry.scaNum), 
				database.countBetSCAQty(entry.scaNum), database.countBetSCA(entry.scaNum), database.countUncheck(entry.scaNum),
				database.currentDate(entry.scaNum))) 
	def clear(self):
		"""Clears tree entries"""
		for entry in self.tree.get_children():
			self.tree.delete(entry)
	def showNSN(self):
		"""Displays count statistics into the treeview"""
		self.entryNSN.delete(first=0, last=100)
		for entry in self.treeNSN.get_children():
			self.treeNSN.delete(entry)
		query = database.displayNSN()
		for entry in query:
			self.treeNSN.insert('','end', text='', value=(entry.itemDes, entry.itemTyp, entry.itemNSN))
	def treeSortCol(self, tv, col, reverse):
		"""Sorts Values Ascending/Descending"""
		l = [(tv.set(k, col), k) for k in tv.get_children('')]
		l.sort(reverse=reverse)
		for index, (val, k) in enumerate(l):
			tv.move(k, '', index)
		tv.heading(col, command=lambda: self.treeSortCol(tv, col, not reverse))
	def selectItem(self, event):
		"""Inserts NSN value into entry box"""
		self.entryNSN.delete(first=0, last=100)
		# Highlights the selected row
		currentItem = self.treeNSN.focus()
		contents = self.treeNSN.item(currentItem)
		values = contents['values']
		self.entryNSN.insert(END, values[2])
################################################################
#  						               #
#                  Page Class - Scan Page                      #
#                                                              #
################################################################
class ScanPage(tk.Frame):
	def __init__(self, parent, child):
		"""Scan and creates a track record for an entry value"""
		tk.Frame.__init__(self, parent)
		#<---------------------------------------------------------------------------| LabelFrame for Scan Equipment
		labelFrameDate = LabelFrame(self, text="Scan Equipment", font=style)
		labelFrameDate.pack(expand=1, fill='both')
		# Label for the display date
		self.date = "Date: " + datetime.datetime.now().strftime('%d-%b-%y')
		tk.Label(labelFrameDate, text=self.date).grid(row=0, column=0, padx=5, sticky='nw')
		#<---------------------------------------------------------------------------| String Value
		self.entryValue = tk.StringVar()
		#<---------------------------------------------------------------------------| Labels and Entry Value
		self.entry = Entry(labelFrameDate, textvariable=self.entryValue, width=40)
		self.entry.bind('<Return>', self.searchFunc)
		self.entry.grid(row=2, column=0, padx=5, pady=5, sticky='w')
		#<---------------------------------------------------------------------------| Button
		tk.Button(labelFrameDate, text="Clear Scan List", command=self.clearListButton, width=20).grid(row=2, column=1, sticky='w')
		#<---------------------------------------------------------------------------| LabelFrame for Treeview and Display Treeview
		labelFrameTree = LabelFrame(self, text="Equipment List", font=style)
		labelFrameTree.pack(expand=1, fill='both')
		self.displayTree = DisplayTreeView(labelFrameTree, "Edit Equipment")
		#<---------------------------------------------------------------------------| Insert Value into an entry using double click function
		self.treeview = self.displayTree.treeView()
		self.treeview.bind('<Double-1>', self.insertValue)
		#<---------------------------------------------------------------------------| LabelFrame for Entry Box
		labelFrameEdit=LabelFrame(self,text="Edit Item", font=style)
		labelFrameEdit.pack(expand=1, fill='both')
		#<---------------------------------------------------------------------------| String Values
		self.valuesID = 0
		self.stringVal = {}
		for i in range(0,10):
			self.stringVal[i] = tk.StringVar()
		#<---------------------------------------------------------------------------| Label and Entry Box
		# Location
		tk.Label(labelFrameEdit, text="Location").grid(row=2, column=0, sticky='w')
		self.entryLoc = Entry(labelFrameEdit, textvariable=self.stringVal[0], width=30)
		self.entryLoc.grid(row=4, column=0, sticky='w')
		# Type
		tk.Label(labelFrameEdit, text="Type").grid(row=2, column=1, sticky='w')
		self.entryTyp = Entry(labelFrameEdit, textvariable=self.stringVal[1], width=18)
		self.entryTyp.grid(row=4, column=1, sticky='w')
		# NSN
		tk.Label(labelFrameEdit, text="NSN").grid(row=2, column=2, sticky='w')
		self.entryNSN = Entry(labelFrameEdit, textvariable=self.stringVal[2], width=20)
		self.entryNSN.grid(row=4, column=2, sticky='w')
		# Model
		tk.Label(labelFrameEdit, text="Model").grid(row=2, column=3, sticky='w')
		self.entryMod = Entry(labelFrameEdit, textvariable=self.stringVal[3], width=20)
		self.entryMod.grid(row=4, column=3, sticky='w')
		# Description
		tk.Label(labelFrameEdit,text="Description").grid(row=2, column=4, sticky='w')
		self.entryDes = Entry(labelFrameEdit, textvariable=self.stringVal[4], width=45)
		self.entryDes.grid(row=4, column=4, sticky='w')
		# Serial Number
		tk.Label(labelFrameEdit, text="Serial Number").grid(row=2, column=5, sticky='w')
		self.entrySer = Entry(labelFrameEdit, textvariable=self.stringVal[5], width=30)
		self.entrySer.grid(row=4, column=5, sticky='w')
		# Asset Number
		tk.Label(labelFrameEdit, text="Asset Number").grid(row=2, column=6, sticky='w')
		self.entryAss = Entry(labelFrameEdit, textvariable=self.stringVal[6], width=20)
		self.entryAss.grid(row=4, column=6, sticky='w')
		# SCA Number
		tk.Label(labelFrameEdit, text="SCA Number").grid(row=2, column=7, sticky='w')
		self.entrySCA = Entry(labelFrameEdit, textvariable=self.stringVal[7], width=20)
		self.entrySCA.grid(row=4, column=7, sticky='w')
		# Tracking Type
		tk.Label(labelFrameEdit, text="Tracking Type").grid(row=2, column=8, sticky='w')
		self.entryTrk = Entry(labelFrameEdit, textvariable=self.stringVal[8], width=20)
		self.entryTrk.grid(row=4, column=8, sticky='w')
		# Quantity
		tk.Label(labelFrameEdit, text="Quantity").grid(row=2, column=9, sticky='w')
		self.entryQty = Entry(labelFrameEdit, textvariable=self.stringVal[9], width=10)
		self.entryQty.grid(row=4, column=9, sticky='w')
		#<---------------------------------------------------------------------------| LabelFrame for Button
		labelFrameAction = LabelFrame(self, text="Action", font=style)
		labelFrameAction.pack(expand=1, fill='both')
		#<---------------------------------------------------------------------------| Button
		tk.Button(labelFrameAction,text="Commit", command=self.commitButton, width=20).grid(row=1, column=0, pady=5, sticky='w')
		tk.Button(labelFrameAction,text="Cancel", command=self.cancelButton, width=20).grid(row=1, column=1, pady=5, sticky='w')
	def cancelButton(self):
		"""Removes all values in entry boxes"""
		self.entryLoc.delete(first=0, last=100)
		self.entryTyp.delete(first=0, last=100)
		self.entryNSN.delete(first=0, last=100)
		self.entryMod.delete(first=0, last=100)
		self.entryDes.delete(first=0, last=100)
		self.entrySer.delete(first=0, last=100)
		self.entryAss.delete(first=0, last=100)
		self.entrySCA.delete(first=0, last=100)
		self.entryTrk.delete(first=0, last=100)
		self.entryQty.delete(first=0, last=100)
	def clearListButton(self):
		"""Clears value in Treeview and entry box"""
		self.displayTree.clearTreeviewButton()
		self.entry.delete(first=0, last=100)
	def commitButton(self):
		"""Create new entry in the Database"""
		#<---------------------------------------------------------------------------| Get String Variable
		ID = self.valuesID
		loc = self.stringVal[0].get()
		typ = self.stringVal[1].get()
		nsn = self.stringVal[2].get()
		mod = self.stringVal[3].get()
		des = self.stringVal[4].get()
		ser = self.stringVal[5].get()
		ass = self.stringVal[6].get()
		sca = self.stringVal[7].get()
		trk = self.stringVal[8].get()
		qty = self.stringVal[9].get()
		# Generate randow number for equipment without serial or asset number
		num = ""
		for i in range(0,9):
			num += str(random.randint(0,9))
		# Assigning random number for equipment with no serial or asset number
		if loc and typ and nsn and mod and des and ser and ass and sca and trk:
			database.update(ID, loc, typ, nsn, mod, des, ser, ass, sca, trk, qty, dte=None)
		elif not ser:
			ser = "<tmp-"+num+">"
			database.update(ID, loc, typ, nsn, mod, des, ser, ass, sca, trk, qty, dte=None)
		elif not ass:
			ass = "<tmp-"+num+">"
			database.update(ID, loc, typ, nsn, mod, des, ser, ass, sca, trk, qty, dte=None)
		# Clears entry list
		self.cancelButton()
		#<---------------------------------------------------------------------------| Returns edited value to the Treeview
		scanID = self.valuesID
		query = database.searchID(scanID)
		self.displayTree.displayTree(query)
		self.cancelButton()
	def insertValue(self, event):
		"""Inserts a string value into an entry box. Executes on double click command"""
		try:
			# Highlights the selected row
			currentItem = self.treeview.focus()
			contents = self.treeview.item(currentItem)
			values = contents['values']
			# Insert tree values into entry box
			self.entryLoc.insert(END, values[0])
			self.entryTyp.insert(END, values[1])
			self.entryNSN.insert(END, values[2])
			self.entryMod.insert(END, values[3])
			self.entryDes.insert(END, values[4])
			self.entrySer.insert(END, values[5])
			self.entryAss.insert(END, values[6])
			self.entrySCA.insert(END, values[7])
			self.entryTrk.insert(END, values[8])
			self.entryQty.insert(END, values[9])
			# Gets the ID number of the treeview content
			contentsID = self.treeview.item(currentItem)
			valuesID = contentsID['text']
			self.valuesID = valuesID
			# Removes the row from the treeview table
			delCurrentItem = self.treeview.focus()
			delContents = self.treeview.delete(delCurrentItem)
		except IndexError:
			pass
	def searchFunc(self, event=None):
		"""Displays value into the treeview. Executes Enter command"""
		value = self.entryValue.get()
		query = database.scan(value)
		self.displayTree.displayTree(query)
		self.entry.delete(first=0, last=100)
################################################################
#  							       #
#                  Page Class - Search Page                    #
#                                                              #
################################################################
class SearchPage(tk.Frame):
	def __init__(self, parent, child):
		"""Searches entry from database"""
		tk.Frame.__init__(self, parent)
		#<---------------------------------------------------------------------------| LabelFrame for Scan Equipment
		labelFrameSearch = LabelFrame(self, text="Search Equipment", font=style)
		labelFrameSearch.pack(expand=1, fill='both')
		#<---------------------------------------------------------------------------| String Values
		self.stringValue = {}
		for i in range(0, 10):
			self.stringValue[i] = tk.StringVar()
		#<---------------------------------------------------------------------------| Label and Entry Box
		# Location
		tk.Label(labelFrameSearch, text="Location").grid(row=2, column=0, sticky='w')
		self.entryLoc = Entry(labelFrameSearch, textvariable=self.stringValue[0], width=40)
		self.entryLoc.bind('<Return>', self.searchButton)
		self.entryLoc.grid(row=4, column=0, sticky='w')
		# Type
		tk.Label(labelFrameSearch, text="Type").grid(row=6, column=0, sticky='w')
		self.entryTyp = Entry(labelFrameSearch, textvariable=self.stringValue[1], width=40)
		self.entryTyp.bind('<Return>', self.searchButton)
		self.entryTyp.grid(row=8, column=0, sticky='w')
		# NSN
		tk.Label(labelFrameSearch, text="NSN <Exact Match>").grid(row=2, column=1, padx=10, pady=5, sticky='w')
		self.entryNSN = Entry(labelFrameSearch, textvariable=self.stringValue[2], width=37)
		self.entryNSN.bind('<Return>', self.searchButton)
		self.entryNSN.grid(row=4, column=1, padx=10, pady=5, sticky='w')
		# Model
		tk.Label(labelFrameSearch, text="Model").grid(row=6, column=1, padx=10, pady=5, sticky='w')
		self.entryMod = Entry(labelFrameSearch, textvariable=self.stringValue[3], width=37)
		self.entryMod.bind('<Return>', self.searchButton)
		self.entryMod.grid(row=8, column=1, padx=10, pady=5, sticky='w')
		# Description
		tk.Label(labelFrameSearch, text="Description").grid(row=2, column=2, sticky='w')
		self.entryDes = Entry(labelFrameSearch, textvariable=self.stringValue[4], width=50)
		self.entryDes.bind('<Return>', self.searchButton)
		self.entryDes.grid(row=4, column=2, sticky='w')
		# Serial Number
		tk.Label(labelFrameSearch, text="Serial Number <Exact Match>").grid(row=6, column=2, sticky='w')
		self.entrySer = Entry(labelFrameSearch, textvariable=self.stringValue[5], width=50)
		self.entrySer.bind('<Return>', self.searchButton)
		self.entrySer.grid(row=8, column=2, sticky='w')
		# Asset Number
		tk.Label(labelFrameSearch, text="Asset Number <Exact Match>").grid(row=2, column=3, padx=10, pady=5, sticky='w')
		self.entryAss = Entry(labelFrameSearch, textvariable=self.stringValue[6], width=37)
		self.entryAss.bind('<Return>', self.searchButton)
		self.entryAss.grid(row=4, column=3, padx=10, pady=5, sticky='w')
		# SCA
		tk.Label(labelFrameSearch, text="SCA Number <Exact Match>").grid(row=6, column=3, padx=10, pady=5, sticky='w')
		self.entrySCA = Entry(labelFrameSearch, textvariable=self.stringValue[7], width=37)
		self.entrySCA.bind('<Return>', self.searchButton)
		self.entrySCA.grid(row=8, column=3, padx=10, pady=5, sticky='w')
		# Tracking Type
		tk.Label(labelFrameSearch, text="Tracking Type").grid(row=2, column=4, sticky='w')
		self.entryTrk = Entry(labelFrameSearch, textvariable=self.stringValue[8], width=37)
		self.entryTrk.bind('<Return>', self.searchButton)
		self.entryTrk.grid(row=4, column=4, sticky='w')
		# Date
		tk.Label(labelFrameSearch, text="Date Last Seen").grid(row=6, column=4, sticky='w')
		self.entryDate = Entry(labelFrameSearch, textvariable=self.stringValue[9], width=37)
		self.entryDate.bind('<Return>', self.searchButton)
		self.entryDate.grid(row=8, column=4, sticky='w')
		#<---------------------------------------------------------------------------| LabelFrame for Button
		tk.Button(labelFrameSearch, text="Search", command=self.searchButton, width=20).grid(row=2, column=6, padx=10, sticky='w')
		#<---------------------------------------------------------------------------| LabelFrame for Treeview and Display Treeview
		labelFrameTree = LabelFrame(self, text="Equipment List", font=style)
		labelFrameTree.pack(expand=1, fill='both')
		self.displayTree = DisplayTreeView(labelFrameTree, "Search Equipment")
		#<---------------------------------------------------------------------------| Passes in Treeview from the Class DisplayTreeView
		self.treeview = self.displayTree.treeView()
		#<---------------------------------------------------------------------------| Button
		tk.Button(labelFrameSearch, text="View All", command=self.viewAllButton, width=20).grid(row=4, column=6, padx=10, sticky='w')
		tk.Button(labelFrameSearch, text="Clear Search List", command=self.clearEntryButton, width=20).grid(row=6, column=6, padx=10, sticky='w')
		tk.Button(labelFrameSearch, text="Clear Equip List", command=self.displayTree.clearTreeviewButton, width=20).grid(row=8, column=6, padx=10, sticky='w')
		tk.Button(labelFrameTree, text="Save", command=self.saveButton, width=20).grid(column=0, row=13, pady=5, sticky='w')
	def clearEntryButton(self):
		"""Delete's values in entry box"""
		listEntry = (self.entryLoc, self.entryTyp, self.entryNSN, self.entryMod, self.entryMod, self.entrySer, self.entryAss, 
			self.entrySCA, self.entryTrk, self.entryDate)
		for cells in listEntry:
			cells.delete(first=0, last=100)
	def saveButton(self):
		"""Saves entry from treeview to Excel Files"""
		try:
			#<---------------------------------------------------------------------------| Opens a dialog box to save a file
			self.filename = filedialog.asksaveasfilename(title="Select File", filetypes=(("Excel Files", "*.xlsx"),("All Files", "*.*")))
			#<---------------------------------------------------------------------------| Creates an Excel File
			wb = Workbook()
			ws = wb.active
			# Tab title
			ws.title = 'Equipment List'
			# Row 1, column title
			ws.append(['Location', 'Type', 'NSN', 'Model', 'Description', 'Serial Number', 
				'Asset Number', 'SCA Number', 'Tracking Type', 'Quantity', 'Date Last Seen'])
			# Inserts values into an Excel file
			entryList = []
			for entry in self.treeview.get_children():
				for values in self.treeview.item(entry, 'values'):
					entryList.append(values)
			if self.filename:
				# Prompts the user File is saved
				messagebox.showinfo("Message", "File Saved")
			# Creates a maximum range from the treeview
			lengthOfList = len(entryList)//11
			for i in range(0,lengthOfList+1):
				loc = entryList[i*11]
				typ = entryList[i*11+1]
				nsn = entryList[i*11+2]
				mod = entryList[i*11+3]
				des = entryList[i*11+4]
				ser = entryList[i*11+5]
				ass = entryList[i*11+6]
				sca = entryList[i*11+7]
				trk = entryList[i*11+8]
				qty = entryList[i*11+9]
				dte = entryList[i*11+10]
				# Passess in the variables to the Excel spreadsheet
				ws.append([loc, typ, nsn, mod, des, ser, ass, sca, trk, qty, dte])
				# Sets the column dimension and header styles
				for i in range(0,11):
					ws.cell(row=1, column=i+1).font = Font(bold=True)
				ws.column_dimensions['A'].width = 24
				ws.column_dimensions['B'].width = 15
				ws.column_dimensions['C'].width = 15
				ws.column_dimensions['D'].width = 15
				ws.column_dimensions['E'].width = 40
				ws.column_dimensions['F'].width = 30
				ws.column_dimensions['G'].width = 20
				ws.column_dimensions['H'].width = 15
				ws.column_dimensions['I'].width = 15
				ws.column_dimensions['J'].width = 10
				ws.column_dimensions['K'].width = 15
				wb.save(self.filename+'.xlsx')
		except IndexError:
			pass
	def searchButton(self, event=None):
		"""Search queries. Executes with Enter command"""
		# Get values from string variable
		loc = self.stringValue[0].get()
		typ = self.stringValue[1].get()
		nsn = self.stringValue[2].get()
		mod = self.stringValue[3].get()
		des = self.stringValue[4].get()
		ser = self.stringValue[5].get()
		ass = self.stringValue[6].get()
		sca = self.stringValue[7].get()
		trk = self.stringValue[8].get()
		dte = self.stringValue[9].get()
		#<---------------------------------------------------------------------------| Search queries to the database
		#Location Query
		if loc and typ and nsn and mod and des and ser and ass and sca and trk and dte:
			date = datetime.datetime.strptime(dte, '%d-%b-%y')
			query = database.searchLTNMDSAScTrDt(loc, typ, nsn, mod, des, ser, ass, sca, trk, date)
		elif loc and typ and nsn and mod and des and ser and ass and sca and dte:
			date = datetime.datetime.strptime(dte, '%d-%b-%y')
			query = database.searchLTNMDSAScDt(loc, typ, nsn, mod, des, ser, ass, sca, date)
		elif loc and typ and nsn and mod and des and ser and ass and sca and trk:
			query = database.searchLTNMDSAScTr(loc, typ, nsn, mod, des, ser, ass, sca, trk)
		elif loc and typ and nsn and mod and des and ser and ass and dte:
			date = datetime.datetime.strptime(dte, '%d-%b-%y')
			query = database.searchLTNMDSADt(loc, typ, nsn, mod, des, ser, ass, date)
		elif loc and typ and nsn and mod and des and ser and ass and trk:
			query = database.searchLTNMDSATr(loc, typ, nsn, mod, des, ser, ass, trk)
		elif loc and typ and nsn and mod and des and ser and ass and sca:
			query = database.searchLTNMDSASc(loc, typ, nsn, mod, des, ser, ass, sca)
		elif loc and typ and nsn and mod and des and ser and dte:
			date = datetime.datetime.strptime(dte, '%d-%b-%y')
			query = database.searchLTNMDSDt(loc, typ, nsn, mod, des, ser, date)
		elif loc and typ and nsn and mod and des and ser and trk:
			query = database.searchLTNMDSTr(loc, typ, nsn, mod, des, ser, trk)
		elif loc and typ and nsn and mod and des and ser and sca:
			query = database.searchLTNMDSSc(loc, typ, nsn, mod, des, ser, sca)
		elif loc and typ and nsn and mod and des and ser and ass:
			query = database.searchLTNMDSA(loc, typ, nsn, mod, des, ser, ass)
		elif loc and typ and nsn and mod and des and dte:
			date = datetime.datetime.strptime(dte, '%d-%b-%y') 	
			query = database.searchLTNMDt(loc, typ, nsn, mod, des, date)
		elif loc and typ and nsn and mod and des and trk:
			query = database.searchLTNMDTr(loc, typ, nsn, mod, des, trk)
		elif loc and typ and nsn and mod and des and sca:
			query = database.searchLTNMDSc(loc, typ, nsn, mod, des, sca)
		elif loc and typ and nsn and mod and des and ass:
			query = database.searchLTNMDA(loc, typ, nsn, mod, des, ass)
		elif loc and typ and nsn and mod and des and ser:
			query = database.searchLTNMDS(loc, typ, nsn, mod, des, ser)
		elif loc and typ and nsn and mod and dte:
			date = datetime.datetime.strptime(dte, '%d-%b-%y')
			query = database.searchLTNMDt(loc, typ, nsn, mod, date)
		elif loc and typ and nsn and mod and trk:
			query = database.searchLTNMTr(loc, typ, nsn, mod, trk)
		elif loc and typ and nsn and mod and sca:
			query = database.searchLTNMSc(loc, typ, nsn, mod, sca)
		elif loc and typ and nsn and mod and ass:
			query = database.searchLTNMA(loc, typ, nsn, mod, ass)
		elif loc and typ and nsn and mod and ser:
			query = database.searchLTNMS(loc, typ, nsn, mod, ser)
		elif loc and typ and nsn and mod and des:
			query = database.searchLTNMD(loc, typ, nsn, mod, des)
		elif loc and typ and nsn and dte:
			date = datetime.datetime.strptime(dte, '%d-%b-%y')
			query = database.searchLTNDt(loc, typ, nsn, date)
		elif loc and typ and nsn and trk:
			query = database.searchLTNTr(loc, typ, nsn, trk)
		elif loc and typ and nsn and sca:
			query = database.searchLTNSc(loc, typ, nsn, sca)
		elif loc and typ and nsn and ass:
			query = database.searchLTNA(loc, typ, nsn, ass)
		elif loc and typ and nsn and ser:
			query = database.searchLTNS(loc, typ, nsn, ser)
		elif loc and typ and nsn and des:
			query = database.searchLTND(loc, typ, nsn, des)
		elif loc and typ and nsn and mod:
			query = database.searchLTNM(loc, typ, nsn, mod)
		elif loc and typ and dte:
			date = datetime.datetime.strptime(dte, '%d-%b-%y')
			query = database.searchLTDt(loc, typ, date)
		elif loc and typ and trk:
			query = database.searchLTTr(loc, typ, trk)
		elif loc and typ and sca:
			query = database.searchLTSc(loc, typ, sca)
		elif loc and typ and ass:
			query = database.searchLTA(loc, typ, ass)
		elif loc and typ and ser:
			query = database.searchLTS(loc, typ, ser)
		elif loc and typ and des:
			query = database.searchLTD(loc, typ, des)
		elif loc and typ and mod:
			query = database.searchLTM(loc, typ, mod)
		elif loc and typ and nsn:
			query = database.searchLTN(loc, typ, nsn)
		elif loc and dte:
			date = datetime.datetime.strptime(dte, '%d-%b-%y')
			query = database.searchLDt(loc, date)
		elif loc and trk:
			query = database.searchLTr(loc, trk)
		elif loc and sca:
			query = database.searchLSc(loc, sca)
		elif loc and ass:
			query = database.searchLA(loc, ass)
		elif loc and ser:
			query = database.searchLS(loc, ser)
		elif loc and des:
			query = database.searchLD(loc, des)
		elif loc and mod:
			query = database.searchLM(loc, mod)
		elif loc and nsn:
			query = database.searchLN(loc, nsn)
		elif loc and typ:
			query = database.searchLT(loc, typ)
		elif loc:
			query = database.searchLoc(loc)
		# Type Query
		elif typ and nsn and mod and des and ser and ass and sca and trk and dte:
			date = datetime.datetime.strptime(dte, '%d-%b-%y')
			query = database.searchTNMDSAScTrDt(typ, nsn, mod, des, ser, ass, sca, trk, date)
		elif typ and nsn and mod and des and ser and ass and sca and dte:
			date = datetime.datetime.strptime(dte, '%d-%b-%y')
			query = database.searchTNMDSAScDt(typ, nsn, mod, des, ser, ass, sca, date)
		elif typ and nsn and mod and des and ser and ass and sca and trk:
			query = database.searchTNMDSAScTr(typ, nsn, mod, des, ser, ass, sca, trk)
		elif typ and nsn and mod and des and ser and ass and dte:
			date = datetime.datetime.strptime(dte, '%d-%b-%y')
			query = database.searchTNMDSADt(typ, nsn, mod, des, ser, ass, date)
		elif typ and nsn and mod and des and ser and ass and trk:
			query = database.searchTNMDSATr(typ, nsn, mod, des, ser, ass, trk)
		elif typ and nsn and mod and des and ser and ass and sca:
			query = database.searchTNMDSASc(typ, nsn, mod, des, ser, ass, sca)
		elif typ and nsn and mod and des and ser and dte:
			date = datetime.datetime.strptime(dte, '%d-%b-%y')
			query = database.searchTNMDSDt(typ, nsn, mod, des, ser, date)
		elif typ and nsn and mod and des and ser and trk:
			query = database.searchTNMDSTr(typ, nsn, mod, des, ser, trk)
		elif typ and nsn and mod and des and ser and sca:
			query = database.searchTNMDSSc(typ, nsn, mod, des, ser, sca)
		elif typ and nsn and mod and des and ser and ass:
			query = database.searchTNMDSA(typ, nsn, mod, des, ser, ass)
		elif typ and nsn and mod and des and dte:
			date = datetime.datetime.strptime(dte, '%d-%b-%y') 	
			query = database.searchTNMDt(typ, nsn, mod, des, date)
		elif typ and nsn and mod and des and trk:
			query = database.searchTNMDTr(typ, nsn, mod, des, trk)
		elif typ and nsn and mod and des and sca:
			query = database.searchTNMDSc(typ, nsn, mod, des, sca)
		elif typ and nsn and mod and des and ass:
			query = database.searchTNMDA(typ, nsn, mod, des, ass)
		elif typ and nsn and mod and des and ser:
			query = database.searchTNMDS(typ, nsn, mod, des, ser)
		elif typ and nsn and mod and dte:
			date = datetime.datetime.strptime(dte, '%d-%b-%y')
			query = database.searchTNMDt(typ, nsn, mod, date)
		elif typ and nsn and mod and trk:
			query = database.searchTNMTr(typ, nsn, mod, trk)
		elif typ and nsn and mod and sca:
			query = database.searchTNMSc(typ, nsn, mod, sca)
		elif typ and nsn and mod and ass:
			query = database.searchTNMA(typ, nsn, mod, ass)
		elif typ and nsn and mod and ser:
			query = database.searchTNMS(typ, nsn, mod, ser)
		elif typ and nsn and mod and des:
			query = database.searchTNMD(typ, nsn, mod, des)
		elif typ and nsn and dte:
			date = datetime.datetime.strptime(dte, '%d-%b-%y')
			query = database.searchTNDt(typ, nsn, date)
		elif typ and nsn and trk:
			query = database.searchTNTr(typ, nsn, trk)
		elif typ and nsn and sca:
			query = database.searchTNSc(typ, nsn, sca)
		elif typ and nsn and ass:
			query = database.searchTNA(typ, nsn, ass)
		elif typ and nsn and ser:
			query = database.searchTNS(typ, nsn, ser)
		elif typ and nsn and des:
			query = database.searchTND(typ, nsn, des)
		elif typ and nsn and mod:
			query = database.searchTNM(typ, nsn, mod)
		elif typ and dte:
			date = datetime.datetime.strptime(dte, '%d-%b-%y')
			query = database.searchTDt(typ, date)
		elif typ and trk:
			query = database.searchTTr(typ, trk)
		elif typ and sca:
			query = database.searchTSc(typ, sca)
		elif typ and ass:
			query = database.searchTA(typ, ass)
		elif typ and ser:
			query = database.searchTS(typ, ser)
		elif typ and des:
			query = database.searchTD(typ, des)
		elif typ and mod:
			query = database.searchTM(typ, mod)
		elif typ and nsn:
			query = database.searchTN(typ, nsn)
		elif typ and mod:
			query= database.searchTyp(typ)
		elif typ:
			query= database.searchTyp(typ)
		# NSN Query
		elif nsn and mod and des and ser and ass and sca and trk and dte:
			date = datetime.datetime.strptime(dte, '%d-%b-%y')
			query = database.searchNMDSAScTrDt(nsn, mod, des, ser, ass, sca, trk, date)
		elif nsn and mod and des and ser and ass and sca and dte:
			date = datetime.datetime.strptime(dte, '%d-%b-%y')
			query = database.searchNMDSAScDt(nsn, mod, des, ser, ass, sca, date)
		elif nsn and mod and des and ser and ass and sca and trk:
			query = database.searchNMDSAScTr(nsn, mod, des, ser, ass, sca, trk)
		elif nsn and mod and des and ser and ass and dte:
			date = datetime.datetime.strptime(dte, '%d-%b-%y')
			query = database.searchNMDSADt(nsn, mod, des, ser, ass, date)
		elif nsn and mod and des and ser and ass and trk:
			query = database.searchNMDSATr(nsn, mod, des, ser, ass, trk)
		elif nsn and mod and des and ser and ass and sca:
			query = database.searchNMDSASc(nsn, mod, des, ser, ass, sca)
		elif nsn and mod and des and ser and dte:
			date = datetime.datetime.strptime(dte, '%d-%b-%y')
			query = database.searchNMDSDt(nsn, mod, des, ser, date)
		elif nsn and mod and des and ser and trk:
			query = database.searchNMDSTr(nsn, mod, des, ser, trk)
		elif nsn and mod and des and ser and sca:
			query = database.searchNMDSSc(nsn, mod, des, ser, sca)
		elif nsn and mod and des and ser and ass:
			query = database.searchNMDSA(nsn, mod, des, ser, ass)
		elif nsn and mod and des and dte:
			date = datetime.datetime.strptime(dte, '%d-%b-%y') 	
			query = database.searchNMDt(nsn, mod, des, date)
		elif nsn and mod and des and trk:
			query = database.searchNMDTr(nsn, mod, des, trk)
		elif nsn and mod and des and sca:
			query = database.searchNMDSc(nsn, mod, des, sca)
		elif nsn and mod and des and ass:
			query = database.searchNMDA(nsn, mod, des, ass)
		elif nsn and mod and des and ser:
			query = database.searchNMDS(nsn, mod, des, ser)
		elif nsn and mod and dte:
			date = datetime.datetime.strptime(dte, '%d-%b-%y')
			query = database.searchNMDt(nsn, mod, date)
		elif nsn and mod and trk:
			query = database.searchNMTr(nsn, mod, trk)
		elif nsn and mod and sca:
			query = database.searchNMSc(nsn, mod, sca)
		elif nsn and mod and ass:
			query = database.searchNMA(nsn, mod, ass)
		elif nsn and mod and ser:
			query = database.searchNMS(nsn, mod, ser)
		elif nsn and mod and des:
			query = database.searchNMD(nsn, mod, des)
		elif nsn and dte:
			date = datetime.datetime.strptime(dte, '%d-%b-%y')
			query = database.searchNDt(nsn, date)
		elif nsn and trk:
			query = database.searchNTr(nsn, trk)
		elif nsn and sca:
			query = database.searchNSc(nsn, sca)
		elif nsn and ass:
			query = database.searchNA(nsn, ass)
		elif nsn and ser:
			query = database.searchNS(nsn, ser)
		elif nsn and des:
			query = database.searchND(nsn, des)
		elif nsn and mod:
			query = database.searchNM(nsn, mod)
		elif nsn:
			query= database.searchNSN(nsn)
		# Model Query
		elif mod and des and ser and ass and sca and trk and dte:
			date = datetime.datetime.strptime(dte, '%d-%b-%y')
			query = database.searchMDSAScTrDt(mod, des, ser, ass, sca, trk, date)
		elif mod and des and ser and ass and sca and dte:
			date = datetime.datetime.strptime(dte, '%d-%b-%y')
			query = database.searchMDSAScDt(mod, des, ser, ass, sca, date)
		elif mod and des and ser and ass and sca and trk:
			query = database.searchMDSAScTr(mod, des, ser, ass, sca, trk)
		elif mod and des and ser and ass and dte:
			date = datetime.datetime.strptime(dte, '%d-%b-%y')
			query = database.searchMDSADt(mod, des, ser, ass, date)
		elif mod and des and ser and ass and trk:
			query = database.searchMDSATr(mod, des, ser, ass, trk)
		elif mod and des and ser and ass and sca:
			query = database.searchMDSASc(mod, des, ser, ass, sca)
		elif mod and des and ser and dte:
			date = datetime.datetime.strptime(dte, '%d-%b-%y')
			query = database.searchMDSDt(mod, des, ser, date)
		elif mod and des and ser and trk:
			query = database.searchMDSTr(mod, des, ser, trk)
		elif mod and des and ser and sca:
			query = database.searchMDSSc(mod, des, ser, sca)
		elif mod and des and ser and ass:
			query = database.searchMDSA(mod, des, ser, ass)
		elif mod and des and dte:
			date = datetime.datetime.strptime(dte, '%d-%b-%y') 	
			query = database.searchMDt(mod, des, date)
		elif mod and des and trk:
			query = database.searchMDTr(mod, des, trk)
		elif mod and des and sca:
			query = database.searchMDSc(mod, des, sca)
		elif mod and des and ass:
			query = database.searchMDA(mod, des, ass)
		elif mod and des and ser:
			query = database.searchMDS(mod, des, ser)
		elif mod and dte:
			date = datetime.datetime.strptime(dte, '%d-%b-%y')
			query = database.searchMDt(mod, date)
		elif mod and trk:
			query = database.searchMTr(mod, trk)
		elif mod and sca:
			query = database.searchMSc(mod, sca)
		elif mod and ass:
			query = database.searchMA(mod, ass)
		elif mod and ser:
			query = database.searchMS(mod, ser)
		elif mod and des:
			query = database.searchMD(mod, des)
		elif mod:
			query= database.searchMod(mod)
		# Desciption Query
		elif des and ser and ass and sca and trk and dte:
			date = datetime.datetime.strptime(dte, '%d-%b-%y')
			query = database.searchDSAScTrDt(des, ser, ass, sca, trk, date)
		elif des and ser and ass and sca and dte:
			date = datetime.datetime.strptime(dte, '%d-%b-%y')
			query = database.searchDSAScDt(des, ser, ass, sca, date)
		elif des and ser and ass and sca and trk:
			query = database.searchDSAScTr(des, ser, ass, sca, trk)
		elif des and ser and ass and dte:
			date = datetime.datetime.strptime(dte, '%d-%b-%y')
			query = database.searchDSADt(des, ser, ass, date)
		elif des and ser and ass and trk:
			query = database.searchDSATr(des, ser, ass, trk)
		elif des and ser and ass and sca:
			query = database.searchDSASc(des, ser, ass, sca)
		elif des and ser and dte:
			date = datetime.datetime.strptime(dte, '%d-%b-%y')
			query = database.searchDSDt(des, ser, date)
		elif des and ser and trk:
			query = database.searchDSTr(des, ser, trk)
		elif des and ser and sca:
			query = database.searchDSSc(des, ser, sca)
		elif des and ser and ass:
			query = database.searchDSA(des, ser, ass)
		elif des and dte:
			date = datetime.datetime.strptime(dte, '%d-%b-%y') 	
			query = database.searchDDt(des, date)
		elif des and trk:
			query = database.searchDTr(des, trk)
		elif des and sca:
			query = database.searchDSc(des, sca)
		elif des and ass:
			query = database.searchDA(des, ass)
		elif des and ser:
			query = database.searchDS(des, ser)
		elif des:
			query= database.searchDes(des)
		# Serial Query
		elif ser and ass and sca and trk and dte:
			date = datetime.datetime.strptime(dte, '%d-%b-%y')
			query = database.searchSAScTrDt(ser, ass, sca, trk, date)
		elif ser and ass and sca and dte:
			date = datetime.datetime.strptime(dte, '%d-%b-%y')
			query = database.searchSAScDt(ser, ass, sca, date)
		elif ser and ass and sca and trk:
			query = database.searchSAScTr(ser, ass, sca, trk)
		elif ser and ass and dte:
			date = datetime.datetime.strptime(dte, '%d-%b-%y')
			query = database.searchSADt(ser, ass, date)
		elif ser and ass and trk:
			query = database.searchSATr(ser, ass, trk)
		elif ser and ass and sca:
			query = database.searchSASc(ser, ass, sca)
		elif ser and dte:
			date = datetime.datetime.strptime(dte, '%d-%b-%y')
			query = database.searchSDt(ser, date)
		elif ser and trk:
			query = database.searchSTr(ser, trk)
		elif ser and sca:
			query = database.searchSSc(ser, sca)
		elif ser and ass:
			query = database.searchDSA(ser, ass)
		elif ser:
			query= database.searchSer(ser)
		# Asset Query
		elif ass and sca and trk and dte:
			date = datetime.datetime.strptime(dte, '%d-%b-%y')
			query = database.searchAScTrDt(ass, sca, trk, date)
		elif ass and sca and dte:
			date = datetime.datetime.strptime(dte, '%d-%b-%y')
			query = database.searchAScDt(ass, sca, date)
		elif ass and sca and trk:
			query = database.searchAScTr(ass, sca, trk)
		elif ass and dte:
			date = datetime.datetime.strptime(dte, '%d-%b-%y')
			query = database.searchADt(ass, date)
		elif ass and trk:
			query = database.searchATr(ass, trk)
		elif ass and sca:
			query = database.searchASc(ass, sca)
		elif ass:
			query= database.searchAss(ass)
		# SCA Query
		elif sca and trk and dte:
			date = datetime.datetime.strptime(dte, '%d-%b-%y')
			query = database.searchScTrDt(sca, trk, date)
		elif sca and dte:
			date = datetime.datetime.strptime(dte, '%d-%b-%y')
			query = database.searchScDt(sca, date)
		elif sca and trk:
			query = database.searchScTr(sca, trk)
		elif sca:
			query = database.searchSCA(sca)
		# Tracking Query
		elif trk and dte:
			date = datetime.datetime.strptime(dte, '%d-%b-%y')
			query = database.searchTrDt(trk, date)
		elif trk:
			query = database.searchTrk(trk)
		# Tracking Query
		elif dte:
			date = datetime.datetime.strptime(dte, '%d-%b-%y')
			query = database.searchDte(date)
		#<---------------------------------------------------------------------------| Display's queries into the Treeview
		try:
			self.displayTree.displayTree(query)
			self.clearEntryButton()
		except UnboundLocalError:
			pass
	def viewAllButton(self):
		"""Displays all content of the database"""
		self.displayTree.clearTreeviewButton()
		query = database.viewAll()
		self.displayTree.displayTree(query)
################################################################
#  							       #
#                 Page Class - Track Page                      #
#                                                              #
################################################################
class TrackPage(tk.Frame):
	def __init__(self, parent, child):
		"""Opens the Track database to track data"""
		tk.Frame.__init__(self, parent)
		#<---------------------------------------------------------------------------| LabelFrame for Track Equipment
		labelFrameTrack = LabelFrame(self, text="Track Equipment", font=style)
		labelFrameTrack.pack(expand=1, fill='both')
		#<---------------------------------------------------------------------------| String Value
		self.idValue = StringVar()
		self.desValue = StringVar()
		self.serValue = StringVar()
		self.assValue = StringVar()
		#<---------------------------------------------------------------------------| Label and Entry Box
		# Unique ID
		tk.Label(labelFrameTrack, text="Unique ID").grid(column=0, row=2, sticky='w')
		self.entryID = Entry(labelFrameTrack, textvariable=self.idValue, width=20)
		self.entryID.bind('<Return>', self.addButton)
		self.entryID.grid(column=0, row=4, sticky='w')
		# Description
		tk.Label(labelFrameTrack, text="Description").grid(column=1, row=2, padx=10, sticky='w')
		self.entryDes = Entry(labelFrameTrack, textvariable=self.desValue, width=40)
		self.entryDes.bind('<Return>', self.addButton)
		self.entryDes.grid(column=1, row=4, padx=10, pady=5, sticky='w')
		# Serial Number
		tk.Label(labelFrameTrack, text="Serial Number").grid(column=2, row=2, sticky='w')
		self.entrySer = Entry(labelFrameTrack, textvariable=self.serValue, width=30)
		self.entrySer.bind('<Return>', self.addButton)
		self.entrySer.grid(column=2, row=4, sticky='w')
		# Asset Number
		tk.Label(labelFrameTrack, text="Asset Number").grid(column=3, row=2, padx=10, sticky='w')
		self.entryAss = Entry(labelFrameTrack, textvariable=self.assValue, width=30)
		self.entryAss.bind('<Return>', self.addButton)
		self.entryAss.grid(column=3, row=4, padx=10, pady=5, sticky='w')
		#<---------------------------------------------------------------------------| LabelFrame for Treeview and Display Treeview
		labelFrameTree = LabelFrame(self, text="Equipment List", font=style)
		labelFrameTree.pack(expand=1, fill='both')
		self.displayTree = DisplayTreeView(labelFrameTree, "Track Equipment")
		#<---------------------------------------------------------------------------| Passes in Treeview from the Class DisplayTreeView
		self.treeview = self.displayTree.treeView()
		self.treeview.bind("<Double-1>", self.insertValue)
		#<---------------------------------------------------------------------------| Button
		# Add value into the Treeview
		tk.Button(labelFrameTrack, text="Add", command=self.addButton, width=20).grid(column=4, row=4, padx=10, sticky='w')
		tk.Button(labelFrameTrack, text="Clear Search Item", command=self.clearSearchButton, width=20).grid(column=5, row=4, sticky='w')
		tk.Button(labelFrameTrack, text="Clear List", command=self.clearTreeviewButton, width=20).grid(column=6, row=4, padx=10, sticky='w')
		#<---------------------------------------------------------------------------| LabelFrame for Treeview and Display Treeview
		labelFrameTreeRecord = LabelFrame(self, text="Record List", font=style)
		labelFrameTreeRecord.pack(expand=1, fill='both')
		self.displayTreeRecord = DisplayTreeView(labelFrameTreeRecord, "TrackRecord Equipment")
		#<---------------------------------------------------------------------------| Passes in Treeview from the Class DisplayTreeView
		self.treeviewRecord = self.displayTreeRecord.treeView()
	def addButton(self, event=None):
		"""Add value into a treeview. Executes to enter command"""
		idValue = self.idValue.get()
		serValue = self.serValue.get()
		desValue = self.desValue.get()
		assValue = self.assValue.get()
		#<---------------------------------------------------------------------------| Displays values into Treeview
		if idValue:
			query = database.searchID(idValue)
		elif serValue:
			query = database.searchSer(serValue)
		elif desValue:
			query = database.searchDes(desValue)
		elif assValue:
			query = database.searchAss(assValue)
		# Displays value in treeview and clears entry box
		self.displayTree.displayTree(query)
		self.clearSearchButton()
	def clearSearchButton(self):
		"""Clears all values in the entry box"""
		self.entryID.delete(first=0, last=100)
		self.entryDes.delete(first=0, last=100)
		self.entrySer.delete(first=0, last=100)
		self.entryAss.delete(first=0, last=100)
	def displayTreeRecView(self, query):
		"""Displays TreeView"""
		for entry in query:
			timestamp = entry.timestampTrack.strftime('%d-%b-%y')
			self.treeviewRecord.insert('','end', text=str(entry.id), value=(
				entry.item_location, 
				entry.item_type,
				entry.item_nsn, 
				entry.item_model, 
				entry.item_desc, 
				entry.item_serial, 
				entry.item_asset, 
				entry.item_sca, 
				entry.item_trk, 
				entry.item_quantity,
				timestamp)) 
	def clearTreeviewButton(self):
		"""Removes display in Treeview"""
		for entry in self.treeviewRecord.get_children():
			self.treeviewRecord.delete(entry)
		self.displayTree.clearTreeviewButton()
	def insertValue(self, event):
		"""Inserts a string value into an entry box. Executes on double click command"""
		try:
			# Removes the previous entry
			for entry in self.treeviewRecord.get_children():
				self.treeviewRecord.delete(entry)
			# Highlights the selected row
			currentItem = self.treeview.focus()
			contents = self.treeview.item(currentItem)
			values = contents['values']
			# Displays value into TrackRecord Tree
			query = database.searchTrackDes(values[5])
			self.displayTreeRecView(query)
		except IndexError:
			pass
if __name__ == '__main__':
	style=('Arial',10,'bold')
	app=App()
	app.title("eTracker")
	app.geometry('1442x750')
	app.resizable(width=tk.FALSE,height=tk.FALSE)
	app.mainloop()
